from __future__ import annotations

from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data"
CSV_PATH = DATA_DIR / "MDS_COHORT.csv"
EXCEL_PATH = DATA_DIR / "Cameron_Patients.xlsx"
OUTPUT_PATH = DATA_DIR / "MDS_COHORT_combined.csv"

MISSING_SENTINELS = {"not in panel", "nd", "na", "none", ""}
SHARED_COLS = [
    "SAMPLE_ID",
    "VISIT_NUMBER",
    "SAMPLE_DATE",
    "GENE",
    "cDNA_CHANGE",
    "PROTEIN_CHANGE",
    "VAF",
    "READ_DEPTH",
    "AGE",
    "SEX",
    "source",
]
MERGE_KEYS = ["SAMPLE_ID", "VISIT_NUMBER", "GENE", "cDNA_CHANGE"]
OVERRIDE_COLS = ["SAMPLE_DATE", "PROTEIN_CHANGE", "VAF", "READ_DEPTH", "AGE", "SEX"]


def parse_vaf(val):
    """Convert an Excel VAF cell to a percentage float, or NaN."""
    if val is None:
        return np.nan

    s = str(val).strip()
    if s.lower() in MISSING_SENTINELS:
        return np.nan
    if s.startswith("<"):
        return 0.5

    try:
        numeric = float(s)
    except ValueError:
        return np.nan

    if 0 <= numeric <= 1:
        return numeric * 100
    if 1 < numeric <= 100:
        return numeric

    return np.nan


def parse_depth(val):
    """Convert an Excel depth cell to float, or NaN."""
    if val is None:
        return np.nan

    s = str(val).strip()
    if s.lower() in MISSING_SENTINELS:
        return np.nan

    try:
        return float(s)
    except ValueError:
        return np.nan


def strip_transcript(cdna):
    """Remove transcript prefix: 'NM_012433.3:c.1876A>G' -> 'c.1876A>G'."""
    if cdna is None:
        return None

    s = str(cdna).strip()
    if ":" in s:
        s = s.split(":", 1)[1]
    return s or None


def normalize_string(val):
    if pd.isna(val):
        return np.nan

    s = str(val).strip()
    return s if s else np.nan


def normalize_date(val):
    if val is None or pd.isna(val):
        return np.nan

    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")

    parsed = pd.to_datetime(val, errors="coerce")
    if pd.isna(parsed):
        s = str(val).strip()
        return s if s else np.nan

    return parsed.strftime("%Y-%m-%d")


def clean_csv_numeric(series):
    series = series.replace("NULL", np.nan)
    series = series.apply(lambda x: str(x).replace(",", ".") if pd.notnull(x) else x)
    series = series.apply(lambda x: 0.5 if pd.notnull(x) and str(x).startswith("<") else x)
    return pd.to_numeric(series, errors="coerce")


def finalize_common_columns(df):
    out = df.copy()

    for col in ["SAMPLE_ID", "GENE", "cDNA_CHANGE", "PROTEIN_CHANGE", "SEX"]:
        out[col] = out[col].apply(normalize_string)

    out["SAMPLE_DATE"] = out["SAMPLE_DATE"].apply(normalize_date)
    out["VISIT_NUMBER"] = pd.to_numeric(out["VISIT_NUMBER"], errors="coerce")
    out["VAF"] = pd.to_numeric(out["VAF"], errors="coerce")
    out["READ_DEPTH"] = pd.to_numeric(out["READ_DEPTH"], errors="coerce")
    out["AGE"] = pd.to_numeric(out["AGE"], errors="coerce")

    return out


def load_excel_rows():
    workbook = load_workbook(EXCEL_PATH, read_only=True)
    excel_rows = []

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        if len(rows) < 4:
            print(f"WARNING: sheet {sheet_name} has too few rows, skipping")
            continue

        date_row = rows[2]
        header_row = rows[3]

        timepoints = []
        i = 5
        while i < len(header_row):
            cell = header_row[i]
            if cell is not None and str(cell).strip().lower() == "vaf":
                date_val = date_row[i] if i < len(date_row) else None
                depth_idx = i + 1 if (i + 1) < len(header_row) else None
                timepoints.append((normalize_date(date_val), i, depth_idx))
                i += 2
            else:
                i += 1

        if not timepoints:
            print(f"WARNING: no timepoint columns in sheet {sheet_name}, skipping")
            continue

        visit_numbers = list(range(1, len(timepoints) + 1))

        for row in rows[4:]:
            gene = row[2] if len(row) > 2 else None
            cdna = row[3] if len(row) > 3 else None
            protein = row[4] if len(row) > 4 else None

            if gene is None or str(gene).strip() == "":
                continue

            cdna_clean = strip_transcript(cdna)

            for visit_num, (date_val, vaf_idx, depth_idx) in zip(visit_numbers, timepoints):
                vaf_raw = row[vaf_idx] if vaf_idx is not None and vaf_idx < len(row) else None
                depth_raw = row[depth_idx] if depth_idx is not None and depth_idx < len(row) else None

                excel_rows.append(
                    {
                        "SAMPLE_ID": normalize_string(sheet_name),
                        "VISIT_NUMBER": visit_num,
                        "SAMPLE_DATE": date_val,
                        "GENE": normalize_string(gene),
                        "cDNA_CHANGE": normalize_string(cdna_clean),
                        "PROTEIN_CHANGE": normalize_string(protein),
                        "VAF": parse_vaf(vaf_raw),
                        "READ_DEPTH": parse_depth(depth_raw),
                        "AGE": np.nan,
                        "SEX": np.nan,
                        "source": "excel",
                    }
                )

    excel_df = pd.DataFrame(excel_rows, columns=SHARED_COLS)
    excel_df = finalize_common_columns(excel_df)

    print(
        f"Excel parsed: {excel_df['SAMPLE_ID'].nunique()} participants, {len(excel_df)} rows"
    )
    print(f"  VAF NaN rate: {excel_df['VAF'].isna().mean():.1%}")

    return excel_df


def load_csv_rows():
    csv_df = pd.read_csv(CSV_PATH, delimiter=";", dtype=str)
    csv_df["source"] = "csv"

    for col in ["AGE", "SEX", "PROTEIN_CHANGE", "SAMPLE_DATE"]:
        if col not in csv_df.columns:
            csv_df[col] = np.nan

    csv_df["VAF"] = clean_csv_numeric(csv_df["VAF"])
    csv_df["READ_DEPTH"] = clean_csv_numeric(csv_df["READ_DEPTH"])
    csv_df = finalize_common_columns(csv_df[SHARED_COLS])

    print(f"CSV parsed:   {csv_df['SAMPLE_ID'].nunique()} participants, {len(csv_df)} rows")
    return csv_df


def drop_duplicate_keys(df, label):
    duplicated = df.duplicated(subset=MERGE_KEYS, keep=False)
    n_dupes = int(duplicated.sum())
    if n_dupes:
        print(f"WARNING: {label} has {n_dupes} rows sharing merge keys; keeping last occurrence")
        df = df.drop_duplicates(subset=MERGE_KEYS, keep="last")
    return df


def merge_rowwise(csv_df, excel_df):
    csv_base = drop_duplicate_keys(csv_df.copy(), "CSV")
    excel_base = drop_duplicate_keys(excel_df.copy(), "Excel")

    combined = csv_base.set_index(MERGE_KEYS).copy()
    excel_indexed = excel_base.set_index(MERGE_KEYS)

    overlapping_keys = combined.index.intersection(excel_indexed.index)
    new_keys = excel_indexed.index.difference(combined.index)

    excel_non_null = excel_indexed[OVERRIDE_COLS].notna()
    overridden_cells = int(
        excel_non_null.loc[overlapping_keys].sum().sum() if len(overlapping_keys) else 0
    )

    for col in OVERRIDE_COLS:
        if col in excel_indexed.columns:
            combined[col] = combined[col].where(
                ~excel_indexed[col].reindex(combined.index).notna(),
                excel_indexed[col].reindex(combined.index),
            )

    if len(overlapping_keys):
        combined.loc[overlapping_keys, "source"] = "excel_override"

    if len(new_keys):
        new_rows = excel_indexed.loc[new_keys].copy()
        combined = pd.concat([combined, new_rows], axis=0)

    combined = combined.reset_index()

    combined["VISIT_NUMBER"] = pd.to_numeric(combined["VISIT_NUMBER"], errors="coerce")
    combined = combined.sort_values(
        ["SAMPLE_ID", "VISIT_NUMBER", "GENE", "cDNA_CHANGE"],
        na_position="last",
    ).reset_index(drop=True)

    stats = {
        "overlapping_rows": int(len(overlapping_keys)),
        "excel_only_rows_added": int(len(new_keys)),
        "excel_non_null_cells_applied": overridden_cells,
    }
    return combined, stats


def print_validation(combined, stats):
    print(
        f"\nCombined: {combined['SAMPLE_ID'].nunique()} participants, {len(combined)} rows"
    )
    print(f"Source breakdown:\n{combined['source'].value_counts(dropna=False).to_string()}")
    print(
        "\nMerge summary:"
        f"\n  overlapping CSV/Excel rows: {stats['overlapping_rows']}"
        f"\n  Excel-only rows added: {stats['excel_only_rows_added']}"
        f"\n  non-null Excel cells applied onto CSV rows: {stats['excel_non_null_cells_applied']}"
    )

    nan_vaf = combined[combined["VAF"].isna()]
    if not nan_vaf.empty:
        print(f"\n{len(nan_vaf)} rows with NaN VAF remain after merge:")
        print(nan_vaf[["SAMPLE_ID", "VISIT_NUMBER", "GENE", "cDNA_CHANGE"]].to_string(index=False))

    duplicate_rows = combined[combined.duplicated(subset=MERGE_KEYS, keep=False)]
    if not duplicate_rows.empty:
        print(f"\nWARNING: {len(duplicate_rows)} rows still share merge keys after merge")
        print(duplicate_rows[MERGE_KEYS + ["source"]].to_string(index=False))


def main():
    if not CSV_PATH.exists():
        raise FileNotFoundError(f"Missing CSV input: {CSV_PATH}")
    if not EXCEL_PATH.exists():
        raise FileNotFoundError(f"Missing Excel input: {EXCEL_PATH}")

    excel_df = load_excel_rows()
    csv_df = load_csv_rows()

    combined, stats = merge_rowwise(csv_df, excel_df)
    print_validation(combined, stats)

    combined.to_csv(OUTPUT_PATH, index=False, sep=";")
    print(f"\nSaved -> {OUTPUT_PATH}")


if __name__ == "__main__":
    main()