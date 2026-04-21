"""
diagnose_excel_parse.py
-----------------------
Inspects the raw Excel structure sheet-by-sheet to identify
where data is being lost in the merge script.
"""

import numpy as np
from openpyxl import load_workbook
from datetime import datetime

EXCEL_PATH = '../data/Cameron_Patients.xlsx'
wb = load_workbook(EXCEL_PATH, read_only=True)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    print(f"\n{'='*70}")
    print(f"SHEET: {sheet_name}  ({len(rows)} rows)")
    print(f"{'='*70}")

    if len(rows) < 5:
        print("  ⚠ Too few rows to parse")
        continue

    # Print first 6 rows raw so we can see actual layout
    print("\nRaw rows 0–5:")
    for i, row in enumerate(rows[:6]):
        print(f"  [{i}] {list(row)}")

    date_row   = rows[2]
    header_row = rows[3]

    print(f"\nDate row   (row 2): {list(date_row)}")
    print(f"Header row (row 3): {list(header_row)}")

    # Show what column indices map to what
    print("\nHeader row column index → value:")
    for idx, val in enumerate(header_row):
        if val is not None:
            print(f"  col {idx:>3}: {repr(val)}")

    # Find VAF/depth column pairs
    timepoints = []
    i = 0
    while i < len(header_row):
        cell = header_row[i]
        if cell is not None and str(cell).strip().lower() == 'vaf':
            date_val = date_row[i] if i < len(date_row) else None
            if isinstance(date_val, datetime):
                date_val = date_val.strftime('%Y-%m-%d')
            depth_idx = i + 1 if (i + 1) < len(header_row) else None
            timepoints.append((date_val, i, depth_idx))
        i += 1

    print(f"\nDetected {len(timepoints)} timepoints:")
    for tp in timepoints:
        print(f"  date={tp[0]}, vaf_col={tp[1]}, depth_col={tp[2]}")

    # Count non-empty mutation rows
    mutation_rows = [r for r in rows[4:] if r[2] is not None and str(r[2]).strip() != '']
    print(f"\nNon-empty mutation rows (col 2 = GENE not blank): {len(mutation_rows)}")

    # Show first 3 mutation rows
    print("First 3 mutation rows:")
    for r in mutation_rows[:3]:
        print(f"  {list(r)}")

    # Check for rows where gene exists but VAF cols are all None/ND
    missing_vaf_count = 0
    for r in mutation_rows:
        vafs = [r[tp[1]] for tp in timepoints if tp[1] < len(r)]
        if all(v is None or str(v).strip().lower() in {'nd', 'not in panel', 'na', 'none', ''} for v in vafs):
            missing_vaf_count += 1
    print(f"Mutation rows with ALL timepoints missing/ND/not-in-panel: {missing_vaf_count}")

    # Check if any rows might be getting skipped due to unexpected gene column position
    print(f"\nChecking col 2 (GENE) values for first 10 data rows:")
    for r in rows[4:14]:
        print(f"  col0={repr(r[0] if len(r)>0 else None)}, "
              f"col1={repr(r[1] if len(r)>1 else None)}, "
              f"col2={repr(r[2] if len(r)>2 else None)}, "
              f"col3={repr(r[3] if len(r)>3 else None)}, "
              f"col4={repr(r[4] if len(r)>4 else None)}")

print("\n" + "="*70)
print("DONE")